# dralun818
"""
Python requests 库使用方法
requests 是 Python 中一个用于发送 HTTP 请求的库，它简化了 HTTP 请求的操作，使您可以轻松地与网站和 API 进行交互。
"""
import requests
import re
url="http://www.weather.com.cn/weather/101010100.shtml"
resp=requests.get(url)
resp.encoding='tuf-8'
print(type(resp.text))#resp响应对象，对象名.属性名

"""
<span class="name">三亚</span>
<span class="weather">雷阵雨</span>
<span class="wd">32/25℃</span>
<span class="zs">一般</span>
"""
city=re.findall('<span class="name">([\u4e00-\u9fa5]*)</span>',resp.text)
weather=re.findall('<span class="weather">([\u4e00-\u9fa5]*)</span>',resp.text)
wd=re.findall('<span class="wd">(.*)</span>',resp.text)
zs=re.findall('<span class="zs">([\u4e00-\u9fa5]*)</span>',resp.text)
"""
注意最外面必须使用()，()内使用[],
"""
# print(city)
# print(weather)
# print(wd)
# print(zs)
lst=[]
for a,b,c,d in zip(city,weather,wd,zs):
    # print(a,b,c,d)
    lst.append([a,b,c,d])
"""
景区 天气 气温 旅游指数
三亚 雷阵雨 32/25℃ 一般
九寨沟 多云转阵雨 32/18℃ 适宜
大理 阴转阵雨 27/18℃ 适宜
张家界 多云 35/26℃ 一般
桂林 多云 35/28℃ 一般
青岛 多云 29/25℃ 适宜
"""
url="https://www.baidu.com/img/PCtm_d9c8750bed0b3c7d089fa7d55720d6cf.png"
resp=requests.get(url)
with open('logo.png','wb') as file:
    file.write(resp.content)





"""
openpyxl 常用操作
openpyxl 是 Python 中一个用于读写 Excel 文件的库。
它提供了丰富的功能，可以方便地操作 Excel 文件中的各种数据和格式。

#读取 Excel 文件
from openpyxl import load_workbook
# 读取文件
wb = load_workbook('data.xlsx')
# 获取工作表
ws = wb['Sheet1']  # 默认第一个工作表，也可以通过索引或名称获取
# 获取单元格值
cell = ws['A1']
print(cell.value)  # 输出：单元格 A1 的值
# 获取所有单元格值
for row in ws.iter_rows():
    for cell in row:
        print(cell.value)



#写入 Excel 文件
from openpyxl import Workbook
# 创建工作簿
wb = Workbook()
# 获取工作表
ws = wb.active  # 默认工作表
# 写入单元格值
ws['A1'] = 'Hello, world!'
ws['B2'] = 123
# 保存文件
wb.save('data.xlsx')
"""





"""
pdfplumber解析和处理PDF文件的Python库

with pdfplumber.open("data.pdf") as pdf:
    for page in pdf.pages:
        text = page.extract_text()
        print(text)

import pdfplumber
i=0
with pdfplumber.open("data.pdf") as pdf:
    for page in pdf.pages:
        text = page.extract_text()#可以提取文件内容
        print(text)
        print('-'*10,page.page_number,'-'*10)
        i+=1
        if i>=10:
            break
"""





"""
numpy
  NumPy 是 Python 语言的一个扩展程序库，用于科学计算。
它提供了一个强大的 N 维数组对象 ndarray，用于存储和操作多维数据，
以及大量的数学函数、随机数生成器、线性代数例程、傅里叶变换等功能。
  NumPy 通常与 SciPy（科学 Python）和 Matplotlib（绘图库）一起使用，
这三个库构成了一个强大的科学计算环境，可用于数据分析、机器学习、图像处理等领域。

NumPy 的核心数据结构是 ndarray 对象。ndarray 对象用于存储多维数据。
要创建 ndarray 对象，可以使用 np.array()
arr = np.array([1, 2, 3, 4, 5, 6, 7, 8, 9, 10])
print(type(arr))#<class 'numpy.ndarray'>

arr = np.array([[1, 2, 3, 4],
                 [5, 6, 7, 8],
                 [9, 10, 11, 12]])
                 
arr1 = np.array([1, 2, 3])
arr2 = np.array([4, 5, 6])
result = arr1 + arr2
print(result)  # 输出： [5 7 9]

arr1 = np.array([1, 2, 3])
arr2 = np.array([4, 5, 6])
result = arr1 * arr2
print(result)  # 输出： [4 10 18]

arr1 = np.array([1, 2, 3])
arr2 = np.array([4])
result = arr1 * arr2
print(result)  # 输出： [4 8 12]

arr = np.array([1, 2, 3, 4, 5])
first_element = arr[0]
print(first_element)  # 输出： 1

arr = np.array([1, 2, 3, 4, 5])
last_three_elements = arr[-3:]
print(last_three_elements)  # 输出： [3 4 5]

arr = np.array([1, 2, 3, 4, 5])
shape = arr.shape
print(shape)  # 输出： (5,)

"""
import numpy as np
import matplotlib.pyplot as plt
n1=plt.imread("img.png")
print(n1.shape)#三维数组，最高维度是图像的高，其次是图像的宽，最低维度是RGB三原色
plt.imshow(n1)
plt.show()
n2=np.array([0.299,0.587,0.114])
x=np.dot(n1,n2)
print(x.shape)
plt.imshow(x)
plt.show()
#(183, 275, 3)--->(183, 275)
"""
imshow() 函数具有许多参数，用于控制图像的显示。以下是一些常用参数：
image: 要显示的图像数据。可以是二维或三维数组、PIL 图像对象或 Matplotlib 路径对象。
cmap: 用于控制图像中不同数值所对应的颜色。可以选择内置的颜色映射，如 gray、hot、jet 等，也可以自定义颜色映射。
norm: 用于控制数值的归一化方式。可以选择 Normalize、LogNorm 等归一化方法。
aspect: 控制图像纵横比（aspect ratio）。可以设置为 auto 或一个数字。
interpolation: 插值方法。用于控制图像的平滑程度和细节程度。可以选择 nearest、bilinear、bicubic 等插值方法。
alpha: 图像透明度。取值范围为 0（完全透明）到 1（完全不透明）之间。
origin: 坐标轴原点的位置。可以设置为 upper 或 lower。
extent: 控制显示的数据范围。可以设置为 [xmin, xmax, ymin, ymax]。
vmin 和 vmax: 控制颜色映射的值域范围。





NumPy 中的 dot() 函数用于计算两个数组之间的点积或矩阵乘积。

import numpy as np
a = np.array([1, 2, 3])
b = np.array([4, 5, 6])
dot_product = np.dot(a, b)
print(dot_product)  # 输出： 32


A = np.array([[1, 2],
              [3, 4]])
B = np.array([[5, 6],
              [7, 8]])
matrix_product = np.dot(A, B)
print(matrix_product)  # 输出： [[19 22]
                       [43 50]]


C = np.array([[[1, 2],
                [3, 4]],
              [[5, 6],
               [7, 8]]])
D = np.array([[[9, 10],
                [11, 12]],
              [[13, 14],
               [15, 16]]])
product = np.dot(C, D)
print(product)  # 输出： [[[ 93  98]
                        [135 144]],
                       [[183 192]
                        [255 264]]]

"""
import numpy as np
A = np.array([[[1, 2, 3],
                [4, 5, 6]],
              [[7, 8, 9],
               [10, 11, 12]]])
v = np.array([1,2,3])
c = np.dot(A, v)
print(c,A.shape,v.shape,c.shape)
#[[14 32]
#[50 68]] (2, 2, 3) (3,) (2, 2)





